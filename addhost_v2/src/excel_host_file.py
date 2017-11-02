from .source_host_file  import SourceHostFile
from openpyxl import load_workbook
from .argument_exception import *

class ExcelHostFile(SourceHostFile):
    def __init__(self, argMgr):
        super(ExcelHostFile,self).__init__(argMgr)
        self.ip_host_dict = {}
        print('ExcelHostFile construtor')
        pass

    def open(self):
        sheetname = self.argMgr.get_sheet_name_argument_value()
        self.wb = load_workbook(self.argMgr.get_file_name_argument_value())
        self.sheetnames = self.wb.get_sheet_names()
        if sheetname not in self.sheetnames:
            raise SheetNameArgumentNotFoundException()
        self.ws = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb.active
        print('ExcelHostFile "open()"')

    def readHosts(self):
        lines = []
        line_num = self.argMgr.get_line_num_argument_value()
        addr_col_num = self.argMgr.get_addr_col_num_argument_value()
        host_name_col_num = self.argMgr.get_host_name_col_num_argument_value()
        for row_cell in self.sheet[self.ws.calculate_dimension()]:
            line = []
            for cell in row_cell:
                line.append(cell.value)
            lines.append(line)
        if len(lines) < line_num:
            raise LineNumArgumentOverFlowException()
        if addr_col_num == host_name_col_num:
            raise AddrColNumArgumentIsEqualToHostNameColNumArgumentException()
        begin = line_num -1
        for idx in range(begin, len(lines)):
            line = lines[idx]
            if len(line) < addr_col_num:
                raise IpAddressColumnNumArgumentOverflowException()
            if len(line) < host_name_col_num:
                raise HostNameColumnNumArgumentOverflowException()
            self.ip_host_dict[line[addr_col_num-1]] = [line[host_name_col_num-1]]
        print('ExcelHostFile "read()"')
        print(self.ip_host_dict)
        return self.ip_host_dict

    def close(self):

        print('ExcelHostFile "close()"')

    def __del__(self):
        print('ExcelHostFile.__del__')
        # SourceHostFile.__del__(self)
        pass