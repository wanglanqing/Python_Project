# coding:utf-8
from openpyxl import load_workbook
from .source_host_file import SourceHostFile
from .argument_exception import *
from .commonlog import CommonLog
import logging


class ExcelHostFile(SourceHostFile):
    def __init__(self, argMgr):
        logging.setLoggerClass(CommonLog)
        self.logger = logging.getLogger(__name__)
        super(ExcelHostFile,self).__init__(argMgr)
        self.sheet = None
        print('ExcelHostFile construtor')
        pass

    def open(self):
        sheetname = self.argMgr.get_sheet_name_argument_value()
        workbook = load_workbook(self.argMgr.get_file_name_argument_value())
        sheetnames = workbook.get_sheet_names()
        if sheetname not in sheetnames:
            raise SheetNameArgumentNotFoundException()
        self.sheet = workbook.get_sheet_by_name(sheetname)
        print('ExcelHostFile "open()"')

    def read_source_host_file(self):
        begin = self.argMgr.get_line_num_argument_value()
        max_row = self.sheet.max_row
        if begin > max_row:
            self.logger(LineNumArgumentOverFlowException())
            raise LineNumArgumentOverFlowException()
        addr_col_num = self.argMgr.get_addr_col_num_argument_value() - 1 # cells元组索引，所以需要减1
        if addr_col_num >= self.sheet.max_column:
            self.logger(IpAddressColumnNumArgumentOverflowException())
            raise IpAddressColumnNumArgumentOverflowException()
        host_name_col_num = self.argMgr.get_host_name_col_num_argument_value() - 1 # cells元组索引，所以需要减1
        if host_name_col_num >= self.sheet.max_column:
            self.logger(HostNameColumnNumArgumentOverflowException())
            raise HostNameColumnNumArgumentOverflowException()
        if addr_col_num == host_name_col_num:
            self.logger(AddrColNumArgumentIsEqualToHostNameColNumArgumentException())
            raise AddrColNumArgumentIsEqualToHostNameColNumArgumentException()


        for row in range(begin, max_row + 1):
            cells = self.sheet[row]
            ip_addr = cells[addr_col_num].value
            if not ip_addr:
                continue
            ip_addr.strip()

            host = cells[host_name_col_num].value
            if not host:
                continue
            host.strip()

            host_set = set()
            if ip_addr in self.ip_hosts_dict:
                host_set = self.ip_hosts_dict[ip_addr]
            else:
                self.ip_hosts_dict[ip_addr] = host_set
            host_set.add(host)
            self.host_ip_dict[host] = ip_addr

    def close(self):
        print('ExcelHostFile "close()"')

    def __del__(self):
        print('ExcelHostFile.__del__')
        # SourceHostFile.__del__(self)
        pass